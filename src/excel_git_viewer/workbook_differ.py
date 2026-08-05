from __future__ import annotations

from collections import OrderedDict
from dataclasses import dataclass
from datetime import date, datetime, time
from hashlib import sha256
from io import BytesIO
from threading import Event, Lock
from typing import ClassVar
from zipfile import BadZipFile, ZipFile, is_zipfile

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from excel_git_viewer.models import (
    CellChange,
    CellChangeType,
    CellContext,
    SheetChange,
    WorkbookDiff,
)


class WorkbookReadError(ValueError):
    """Raised when workbook bytes are invalid, unsupported, or unsafe."""


class OperationCancelled(RuntimeError):
    """Raised when the caller cancels an in-flight comparison."""


class CancellationToken:
    def __init__(self) -> None:
        self._event = Event()

    def cancel(self) -> None:
        self._event.set()

    @property
    def is_cancelled(self) -> bool:
        return self._event.is_set()

    def raise_if_cancelled(self) -> None:
        if self.is_cancelled:
            raise OperationCancelled


@dataclass(frozen=True, slots=True)
class _CellSnapshot:
    value: str
    data_type: str
    hidden_sheet: bool
    hidden_row: bool
    hidden_column: bool


@dataclass(frozen=True, slots=True)
class _WorkbookSnapshot:
    cells: dict[tuple[str, str], _CellSnapshot]
    sheets: frozenset[str]


class WorkbookDiffer:
    """Compare supported workbook content through one read-only interface."""

    MAX_ZIP_ENTRIES = 10_000
    MAX_ZIP_ENTRY_BYTES = 64 * 1024 * 1024
    MAX_ZIP_TOTAL_BYTES = 256 * 1024 * 1024
    SNAPSHOT_CACHE_MAX_BYTES = 384 * 1024 * 1024
    _snapshot_cache: ClassVar[OrderedDict[str, tuple[_WorkbookSnapshot, int]]] = OrderedDict()
    _snapshot_cache_bytes: ClassVar[int] = 0
    _snapshot_cache_lock: ClassVar[Lock] = Lock()

    def compare(
        self,
        old_bytes: bytes | None,
        new_bytes: bytes | None,
        *,
        cancellation: CancellationToken | None = None,
    ) -> WorkbookDiff:
        self._raise_if_cancelled(cancellation)
        old_workbook = self._read_workbook(old_bytes, cancellation)
        new_workbook = self._read_workbook(new_bytes, cancellation)
        old_cells = old_workbook.cells
        new_cells = new_workbook.cells
        changes: list[CellChange] = []

        for sheet_name, coordinate in sorted(old_cells.keys() | new_cells.keys()):
            self._raise_if_cancelled(cancellation)
            old_cell = old_cells.get((sheet_name, coordinate))
            new_cell = new_cells.get((sheet_name, coordinate))
            old_value = old_cell.value if old_cell is not None else None
            new_value = new_cell.value if new_cell is not None else None
            old_data_type = old_cell.data_type if old_cell is not None else None
            new_data_type = new_cell.data_type if new_cell is not None else None
            if old_value == new_value and old_data_type == new_data_type:
                continue
            if old_value is None:
                change_type: CellChangeType = "added"
            elif new_value is None:
                change_type = "deleted"
            else:
                change_type = "modified"
            changes.append(
                CellChange(
                    sheet_name=sheet_name,
                    coordinate=coordinate,
                    change_type=change_type,
                    old_value=old_value,
                    new_value=new_value,
                    old_data_type=old_data_type,
                    new_data_type=new_data_type,
                    hidden_sheet=bool(
                        (old_cell and old_cell.hidden_sheet) or (new_cell and new_cell.hidden_sheet)
                    ),
                    hidden_row=bool(
                        (old_cell and old_cell.hidden_row) or (new_cell and new_cell.hidden_row)
                    ),
                    hidden_column=bool(
                        (old_cell and old_cell.hidden_column)
                        or (new_cell and new_cell.hidden_column)
                    ),
                    whitespace_warning=self._has_invisible_whitespace(old_value)
                    or self._has_invisible_whitespace(new_value),
                    old_context=(
                        self._build_context(old_cells, sheet_name, coordinate)
                        if old_bytes is not None
                        else None
                    ),
                    new_context=(
                        self._build_context(new_cells, sheet_name, coordinate)
                        if new_bytes is not None
                        else None
                    ),
                )
            )

        sheet_changes = tuple(
            [
                SheetChange(sheet_name=name, change_type="deleted")
                for name in sorted(old_workbook.sheets - new_workbook.sheets)
            ]
            + [
                SheetChange(sheet_name=name, change_type="added")
                for name in sorted(new_workbook.sheets - old_workbook.sheets)
            ]
        )
        return WorkbookDiff(cell_changes=tuple(changes), sheet_changes=sheet_changes)

    @staticmethod
    def _read_workbook(
        xlsx_bytes: bytes | None,
        cancellation: CancellationToken | None,
    ) -> _WorkbookSnapshot:
        WorkbookDiffer._raise_if_cancelled(cancellation)
        if xlsx_bytes is None:
            return _WorkbookSnapshot(cells={}, sheets=frozenset())
        cache_key = sha256(xlsx_bytes).hexdigest()
        cached = WorkbookDiffer._get_cached_snapshot(cache_key)
        if cached is not None:
            return cached
        WorkbookDiffer._preflight_zip(xlsx_bytes)
        try:
            workbook = load_workbook(
                BytesIO(xlsx_bytes),
                read_only=False,
                data_only=False,
                keep_links=False,
            )
        except Exception as error:
            raise WorkbookReadError("The file is not a valid xlsx workbook") from error
        try:
            cells: dict[tuple[str, str], _CellSnapshot] = {}
            for sheet in workbook.worksheets:
                WorkbookDiffer._raise_if_cancelled(cancellation)
                hidden_columns = WorkbookDiffer._hidden_columns(sheet)
                for row in sheet.iter_rows():
                    WorkbookDiffer._raise_if_cancelled(cancellation)
                    for cell in row:
                        if cell.value is not None:
                            cells[(sheet.title, cell.coordinate)] = _CellSnapshot(
                                value=WorkbookDiffer._normalize_value(cell.value),
                                data_type=WorkbookDiffer._normalize_data_type(cell.data_type),
                                hidden_sheet=sheet.sheet_state != "visible",
                                hidden_row=bool(sheet.row_dimensions[cell.row].hidden),
                                hidden_column=cell.column in hidden_columns,
                            )
            snapshot = _WorkbookSnapshot(cells=cells, sheets=frozenset(workbook.sheetnames))
            WorkbookDiffer._store_cached_snapshot(cache_key, snapshot)
            return snapshot
        finally:
            workbook.close()

    @classmethod
    def clear_snapshot_cache(cls) -> None:
        with cls._snapshot_cache_lock:
            cls._snapshot_cache.clear()
            cls._snapshot_cache_bytes = 0

    @classmethod
    def _get_cached_snapshot(cls, cache_key: str) -> _WorkbookSnapshot | None:
        with cls._snapshot_cache_lock:
            cached = cls._snapshot_cache.get(cache_key)
            if cached is None:
                return None
            cls._snapshot_cache.move_to_end(cache_key)
            return cached[0]

    @classmethod
    def _store_cached_snapshot(cls, cache_key: str, snapshot: _WorkbookSnapshot) -> None:
        estimated_bytes = cls._estimate_snapshot_bytes(snapshot)
        if estimated_bytes > cls.SNAPSHOT_CACHE_MAX_BYTES:
            return
        with cls._snapshot_cache_lock:
            previous = cls._snapshot_cache.pop(cache_key, None)
            if previous is not None:
                cls._snapshot_cache_bytes -= previous[1]
            cls._snapshot_cache[cache_key] = (snapshot, estimated_bytes)
            cls._snapshot_cache_bytes += estimated_bytes
            while cls._snapshot_cache_bytes > cls.SNAPSHOT_CACHE_MAX_BYTES:
                _, (_, removed_bytes) = cls._snapshot_cache.popitem(last=False)
                cls._snapshot_cache_bytes -= removed_bytes

    @staticmethod
    def _estimate_snapshot_bytes(snapshot: _WorkbookSnapshot) -> int:
        estimated_bytes = 512 + sum(128 + len(name) * 2 for name in snapshot.sheets)
        for (sheet_name, coordinate), cell in snapshot.cells.items():
            estimated_bytes += (
                256
                + len(sheet_name) * 2
                + len(coordinate) * 2
                + len(cell.value) * 2
                + len(cell.data_type) * 2
            )
        return estimated_bytes

    @staticmethod
    def _preflight_zip(xlsx_bytes: bytes) -> None:
        if xlsx_bytes.startswith(b"version https://git-lfs.github.com/spec/v1"):
            raise WorkbookReadError(
                "The Git object is a Git LFS pointer; pull the LFS object before reviewing it"
            )
        if xlsx_bytes.startswith(bytes.fromhex("D0CF11E0A1B11AE1")):
            raise WorkbookReadError(
                "Encrypted or legacy Office Compound File workbooks are not supported"
            )
        stream = BytesIO(xlsx_bytes)
        if not is_zipfile(stream):
            raise WorkbookReadError("The file is not a valid xlsx ZIP package")
        try:
            with ZipFile(stream) as archive:
                entries = archive.infolist()
                if len(entries) > WorkbookDiffer.MAX_ZIP_ENTRIES:
                    raise WorkbookReadError("The xlsx package contains too many ZIP entries")
                total_size = 0
                for entry in entries:
                    if entry.flag_bits & 0x1:
                        raise WorkbookReadError("Encrypted workbooks are not supported")
                    if entry.file_size > WorkbookDiffer.MAX_ZIP_ENTRY_BYTES:
                        raise WorkbookReadError("The xlsx package contains an oversized ZIP entry")
                    total_size += entry.file_size
                    if total_size > WorkbookDiffer.MAX_ZIP_TOTAL_BYTES:
                        raise WorkbookReadError("The xlsx package expands beyond the safety limit")
        except BadZipFile as error:
            raise WorkbookReadError("The file is not a valid xlsx ZIP package") from error

    @staticmethod
    def _build_context(
        cells: dict[tuple[str, str], _CellSnapshot],
        sheet_name: str,
        coordinate: str,
        radius: int = 2,
    ) -> CellContext:
        target_row, target_column = coordinate_to_tuple(coordinate)
        start_row = max(1, target_row - radius)
        start_column = max(1, target_column - radius)
        end_row = target_row + radius
        end_column = target_column + radius
        values: list[tuple[str | None, ...]] = []
        for row in range(start_row, end_row + 1):
            row_values: list[str | None] = []
            for column in range(start_column, end_column + 1):
                snapshot = cells.get((sheet_name, WorkbookDiffer._coordinate(row, column)))
                row_values.append(snapshot.value if snapshot is not None else None)
            values.append(tuple(row_values))
        return CellContext(
            start_row=start_row,
            start_column=start_column,
            values=tuple(values),
        )

    @staticmethod
    def _coordinate(row: int, column: int) -> str:
        return f"{get_column_letter(column)}{row}"

    @staticmethod
    def _has_invisible_whitespace(value: str | None) -> bool:
        if not value:
            return False
        return (
            value != value.strip()
            or "\t" in value
            or "\r" in value
            or "\n" in value
            or "  " in value
        )

    @staticmethod
    def _normalize_value(value: object) -> str:
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, datetime | date | time):
            return value.isoformat()
        return str(value)

    @staticmethod
    def _normalize_data_type(data_type: str) -> str:
        return {
            "b": "boolean",
            "d": "date",
            "e": "error",
            "f": "formula",
            "n": "number",
            "s": "text",
            "inlineStr": "text",
            "str": "text",
        }.get(data_type, data_type)

    @staticmethod
    def _hidden_columns(sheet: Worksheet) -> frozenset[int]:
        hidden_columns: set[int] = set()
        for dimension in sheet.column_dimensions.values():
            start = dimension.min or 0
            end = dimension.max or start
            if dimension.hidden:
                hidden_columns.update(range(start, end + 1))
        return frozenset(hidden_columns)

    @staticmethod
    def _raise_if_cancelled(cancellation: CancellationToken | None) -> None:
        if cancellation is not None:
            cancellation.raise_if_cancelled()
