from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path

import pytest
import xlwt  # type: ignore[import-untyped]
from openpyxl import Workbook

import excel_git_viewer.workbook_differ as workbook_differ_module
from excel_git_viewer.workbook_differ import (
    CancellationToken,
    OperationCancelled,
    WorkbookDiffer,
    WorkbookReadError,
)


def workbook_bytes(
    cells: dict[str, object],
    *,
    hidden_sheet: bool = False,
    hidden_rows: set[int] | None = None,
    hidden_columns: set[str] | None = None,
    hidden_column_range: tuple[str, str] | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Items"
    for coordinate, value in cells.items():
        sheet[coordinate] = value
    if hidden_sheet:
        sheet.sheet_state = "hidden"
        workbook.create_sheet("Visible")
    for row in hidden_rows or set():
        sheet.row_dimensions[row].hidden = True
    for column in hidden_columns or set():
        sheet.column_dimensions[column].hidden = True
    if hidden_column_range is not None:
        sheet.column_dimensions.group(*hidden_column_range, hidden=True)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def xls_workbook_bytes(
    cells: dict[str, object],
    *,
    hidden_sheet: bool = False,
    hidden_rows: set[int] | None = None,
    hidden_columns: set[int] | None = None,
) -> bytes:
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Items")
    if hidden_sheet:
        sheet.visibility = 1
    date_style = xlwt.easyxf(num_format_str="YYYY-MM-DD")
    for coordinate, value in cells.items():
        row, column = workbook_differ_module.coordinate_to_tuple(coordinate)
        style = date_style if isinstance(value, date) else xlwt.Style.default_style
        sheet.write(row - 1, column - 1, value, style)
    for row in hidden_rows or set():
        sheet.row(row - 1).hidden = True
    for column in hidden_columns or set():
        sheet.col(column - 1).hidden = True
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_changed_text_cell_reports_its_old_and_new_value() -> None:
    old_workbook = workbook_bytes({"B2": "Potion"})
    new_workbook = workbook_bytes({"B2": "Large Potion"})

    result = WorkbookDiffer().compare(old_workbook, new_workbook)

    assert [
        (
            change.sheet_name,
            change.coordinate,
            change.change_type,
            change.old_value,
            change.new_value,
        )
        for change in result.cell_changes
    ] == [("Items", "B2", "modified", "Potion", "Large Potion")]


def test_formula_diff_compares_formula_text_instead_of_cached_results() -> None:
    old_workbook = workbook_bytes({"D2": "=B2*C2"})
    new_workbook = workbook_bytes({"D2": "=B2*E2"})

    result = WorkbookDiffer().compare(old_workbook, new_workbook)

    assert [(change.old_value, change.new_value) for change in result.cell_changes] == [
        ("=B2*C2", "=B2*E2")
    ]


def test_change_marks_hidden_sheet_row_and_column() -> None:
    old_workbook = workbook_bytes({"B2": "old"})
    new_workbook = workbook_bytes(
        {"B2": "new"},
        hidden_sheet=True,
        hidden_rows={2},
        hidden_columns={"B"},
    )

    [change] = WorkbookDiffer().compare(old_workbook, new_workbook).cell_changes

    assert (change.hidden_sheet, change.hidden_row, change.hidden_column) == (True, True, True)


def test_change_includes_read_only_context_around_the_coordinate() -> None:
    old_workbook = workbook_bytes({"A1": "ID", "B1": "Price", "A3": 1001, "B3": 10})
    new_workbook = workbook_bytes({"A1": "ID", "B1": "Price", "A3": 1001, "B3": 20})

    [change] = WorkbookDiffer().compare(old_workbook, new_workbook).cell_changes

    assert change.old_context is not None
    assert change.old_context.start_row == 1
    assert change.old_context.start_column == 1
    assert change.old_context.values[0][:2] == ("ID", "Price")
    assert change.old_context.values[2][:2] == ("1001", "10")
    assert change.new_context is not None
    assert change.new_context.values[2][:2] == ("1001", "20")


def test_whitespace_is_preserved_and_marked_for_display() -> None:
    old_workbook = workbook_bytes({"A1": "Potion"})
    new_workbook = workbook_bytes({"A1": "Potion "})

    [change] = WorkbookDiffer().compare(old_workbook, new_workbook).cell_changes

    assert (change.old_value, change.new_value, change.whitespace_warning) == (
        "Potion",
        "Potion ",
        True,
    )


def test_empty_added_sheet_is_reported() -> None:
    old_workbook = workbook_bytes({})
    workbook = Workbook()
    workbook.active.title = "Items"
    workbook.create_sheet("New Sheet")
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    result = WorkbookDiffer().compare(old_workbook, output.getvalue())

    assert [(change.sheet_name, change.change_type) for change in result.sheet_changes] == [
        ("New Sheet", "added")
    ]


def test_invalid_workbook_is_reported_as_a_read_error() -> None:
    with pytest.raises(WorkbookReadError, match="valid xlsx"):
        WorkbookDiffer().compare(b"not a zip package", workbook_bytes({}))


def test_cancelled_comparison_stops_before_parsing() -> None:
    token = CancellationToken()
    token.cancel()

    with pytest.raises(OperationCancelled):
        WorkbookDiffer().compare(workbook_bytes({"A1": "old"}), None, cancellation=token)


def test_sanitized_microsoft_excel_fixture_is_readable() -> None:
    fixture = Path(__file__).parent / "fixtures" / "excel" / "blank-excel.xlsx"
    data = fixture.read_bytes()

    result = WorkbookDiffer().compare(data, data)

    assert result.cell_changes == ()


def test_dates_and_booleans_use_stable_normalized_values() -> None:
    old_workbook = workbook_bytes({"A1": date(2026, 8, 4), "B1": True})
    new_workbook = workbook_bytes({"A1": date(2026, 8, 5), "B1": False})

    changes = WorkbookDiffer().compare(old_workbook, new_workbook).cell_changes

    assert [(change.old_value, change.new_value) for change in changes] == [
        ("2026-08-04T00:00:00", "2026-08-05T00:00:00"),
        ("true", "false"),
    ]


def test_xls_cell_changes_use_the_same_normalized_diff_model() -> None:
    old_workbook = xls_workbook_bytes(
        {"A1": "Potion", "B2": 10, "C2": date(2026, 8, 4), "D2": True}
    )
    new_workbook = xls_workbook_bytes(
        {"A1": "Large Potion", "B2": 12, "C2": date(2026, 8, 5), "D2": False},
        hidden_rows={2},
        hidden_columns={2},
    )

    changes = WorkbookDiffer().compare(old_workbook, new_workbook).cell_changes

    assert [
        (change.coordinate, change.old_value, change.new_value, change.new_data_type)
        for change in changes
    ] == [
        ("A1", "Potion", "Large Potion", "text"),
        ("B2", "10", "12", "number"),
        ("C2", "2026-08-04T00:00:00", "2026-08-05T00:00:00", "date"),
        ("D2", "true", "false", "boolean"),
    ]
    assert changes[1].hidden_row is True
    assert changes[1].hidden_column is True


def test_xls_added_and_deleted_cells_and_sheets_are_reported() -> None:
    old_workbook = xls_workbook_bytes({"A1": "removed"})
    new_workbook = xls_workbook_bytes({"B2": "added"})

    result = WorkbookDiffer().compare(old_workbook, new_workbook)

    assert [(change.coordinate, change.change_type) for change in result.cell_changes] == [
        ("A1", "deleted"),
        ("B2", "added"),
    ]


def test_numeric_value_changed_to_equal_text_is_still_reported() -> None:
    old_workbook = workbook_bytes({"A1": 1})
    new_workbook = workbook_bytes({"A1": "1"})

    [change] = WorkbookDiffer().compare(old_workbook, new_workbook).cell_changes

    assert (change.old_value, change.new_value, change.old_data_type, change.new_data_type) == (
        "1",
        "1",
        "number",
        "text",
    )


def test_change_inside_grouped_hidden_columns_is_marked() -> None:
    old_workbook = workbook_bytes({"C2": "old"})
    new_workbook = workbook_bytes(
        {"C2": "new"},
        hidden_column_range=("B", "D"),
    )

    [change] = WorkbookDiffer().compare(old_workbook, new_workbook).cell_changes

    assert change.hidden_column is True


def test_repeated_comparisons_reuse_parsed_workbook_snapshots(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    WorkbookDiffer.clear_snapshot_cache()
    data = workbook_bytes({"A1": "Potion", "B2": 100})
    load_count = 0
    original_load_workbook = workbook_differ_module.load_workbook

    def counted_load_workbook(*args: object, **kwargs: object) -> object:
        nonlocal load_count
        load_count += 1
        return original_load_workbook(*args, **kwargs)

    monkeypatch.setattr(workbook_differ_module, "load_workbook", counted_load_workbook)

    WorkbookDiffer().compare(data, data)
    WorkbookDiffer().compare(data, data)

    assert load_count == 1
    WorkbookDiffer.clear_snapshot_cache()


def test_git_lfs_pointer_has_a_specific_diagnostic() -> None:
    pointer = (
        b"version https://git-lfs.github.com/spec/v1\noid sha256:0123456789abcdef\nsize 12345\n"
    )

    with pytest.raises(WorkbookReadError, match="Git LFS"):
        WorkbookDiffer().compare(pointer, None)


def test_invalid_or_encrypted_legacy_container_has_a_specific_diagnostic() -> None:
    encrypted_container = bytes.fromhex("D0CF11E0A1B11AE1") + b"encrypted payload"

    with pytest.raises(WorkbookReadError, match="valid .xls|encrypted"):
        WorkbookDiffer().compare(encrypted_container, None)
